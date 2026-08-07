"""Excel 台账导入服务：preview（解析+映射）/ confirm（入库+幂等）。
依据 doc/Excel导入映射.md。默认迁"项目/节点/任务骨架 + 当前周目标"。"""
import re
from datetime import date
from typing import Optional

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.project_roles import PROJECT_ROLE_NAMES, canonical_project_role, empty_role_assignments
from app.models.project import Project, ProjectMember, ProjectNode, Task, TrTemplate, TrTemplateNode
from app.models.user import User
from app.services.progress_service import ProgressService
from app.services.project_service import ProjectService

# 节点关键词映射（可配置化）
NODE_KEYWORDS = {
    "TR1": ["tr1", "概念", "需求"],
    "TR2": ["tr2", "方案", "计划"],
    "TR3": ["tr3", "设计", "开发前"],
    "TR4": ["tr4", "开发", "编码", "测试"],
    "TR5": ["tr5", "发布", "验证"],
    "TR6": ["tr6", "收尾", "量产"],
}
OWNER_KEYWORDS = ["负责人", "pm", "owner", "项目经理"]
OWNER_LABELS = {"负责人", "owner", "pm", "项目经理"}


def _match_node_key(text: str) -> Optional[str]:
    if not text:
        return None
    t = str(text).strip().lower()
    for key, kws in NODE_KEYWORDS.items():
        if any(k in t for k in kws):
            return key
    return None


def _split_tasks(text: str) -> list[str]:
    if not text:
        return []
    parts = re.split(r"[\n；;、]|(?:\d+[.、\)])", str(text))
    return [p.strip() for p in parts if p and p.strip()]


def _split_names(text: str) -> list[str]:
    if not text:
        return []
    return [part.strip() for part in re.split(r"[、,，;；\n]+", text) if part and part.strip()]


def _parse_role_cell(text: str) -> tuple[dict[str, list[str]], list[str], bool]:
    """解析“项目角色”单元格，返回固定角色、负责人和是否识别到结构化行。"""
    roles = {role: [] for role in PROJECT_ROLE_NAMES}
    owners = []
    if not text:
        return roles, owners, False

    # Excel 通常传入实际换行，同时兼容用户手工输入的字面量 \\n / \\N。
    normalized = str(text).replace("\\r\\n", "\n").replace("\\n", "\n").replace("\\N", "\n")
    normalized = normalized.replace("\r\n", "\n").replace("\r", "\n")
    detected = False
    for line in normalized.split("\n"):
        line = line.strip()
        if not line:
            continue
        match = re.match(r"^([^:：]+?)\s*[:：]\s*(.*?)\s*$", line)
        if not match:
            continue
        label, names_text = match.groups()
        names = _split_names(names_text)
        role = canonical_project_role(label)
        if role:
            detected = True
            for name in names:
                if name not in roles[role]:
                    roles[role].append(name)
            continue
        if label.strip().casefold() in OWNER_LABELS:
            detected = True
            for name in names:
                if name not in owners:
                    owners.append(name)
    return roles, owners, detected


class ImportService:
    def __init__(self, db: Session):
        self.db = db
        self.ps = ProgressService(db)

    # ---------- preview ----------
    def preview(self, file_bytes: bytes) -> dict:
        wb = load_workbook(filename=__import__("io").BytesIO(file_bytes), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {"projects": [], "warnings": ["空表格"], "errors": []}
        header = [str(c).strip() if c else "" for c in rows[0]]
        col = {name: idx for idx, name in enumerate(header)}

        def gv(row, *names):
            for n in names:
                if n in col and col[n] < len(row) and row[col[n]] is not None:
                    return str(row[col[n]]).strip()
            return ""

        users_by_name = {u.display_name.strip(): u for u in self.db.execute(select(User)).scalars().all()}

        def add_member(p: dict, name: str, user: Optional[User], role: str, invested: bool, is_owner: bool = False) -> None:
            identity = user.id if user else f"name:{name}"
            key = (identity, role)
            for member in p["members"]:
                if member["_key"] == key:
                    member["is_invested"] = invested
                    member["is_owner"] = member["is_owner"] or is_owner
                    return
            p["members"].append({
                "_key": key,
                "name": name, "user_id": user.id if user else None, "matched": bool(user),
                "project_role": role, "is_owner": is_owner, "is_invested": invested,
            })

        def add_role_assignment(p: dict, role: str, name: str, user: Optional[User], invested: bool) -> None:
            entries = p["role_assignments"][role]
            identity = user.id if user else f"name:{name}"
            if not any(item["_key"] == identity for item in entries):
                entries.append({
                    "_key": identity,
                    "name": name, "user_id": user.id if user else None, "matched": bool(user),
                })
            add_member(p, name, user, role, invested)

        projects = {}
        warnings, errors = [], []
        for r in rows[1:]:
            if not any(r):
                continue
            machine = gv(r, "机型")
            pname = gv(r, "项目", "项目名称")
            if not pname:
                warnings.append("存在项目名为空的行，已跳过")
                continue
            key = f"{machine}|{pname}"
            person = gv(r, "成员", "姓名", "人员")
            invested = gv(r, "是否投入该项目", "是否投入")
            role = gv(r, "项目角色")
            role_lines, owner_names, role_cell_detected = _parse_role_cell(role)
            node_text = gv(r, "关键节点", "节点", "当前节点")
            week_goal = gv(r, "项目周目标", "周目标")
            week_task = gv(r, "本周任务", "本周工作")

            p = projects.setdefault(key, {
                "name": pname, "machine_model": machine, "code": "",
                "members": [], "node_text": node_text, "weekly_goal": week_goal,
                "tasks": [], "warnings": [], "role_assignments": empty_role_assignments(),
                "role_assignments_detected": False,
            })
            if week_goal and not p["weekly_goal"]:
                p["weekly_goal"] = week_goal

            if role_cell_detected:
                p["role_assignments_detected"] = True
                role_invested = invested not in ("否", "N", "n", "0") if invested else True
                for role_name, names in role_lines.items():
                    for name in names:
                        add_role_assignment(p, role_name, name, users_by_name.get(name), role_invested)
                for name in owner_names:
                    add_member(p, name, users_by_name.get(name), "负责人", role_invested, is_owner=True)

            # 兼容旧模板：成员姓名单独在“成员/姓名/人员”列中。
            if not person and role and not role_cell_detected:
                if not canonical_project_role(role) and role.strip().casefold() not in OWNER_LABELS:
                    person, role = role, "成员"
            if person:
                u = users_by_name.get(person)
                is_owner = any(k in (role or "").lower() for k in OWNER_KEYWORDS)
                member_role = role or ("负责人" if is_owner else "成员")
                member_invested = invested not in ("否", "N", "n", "0", "")
                add_member(p, person, u, member_role, member_invested, is_owner=is_owner)
                fixed_role = canonical_project_role(role)
                if fixed_role:
                    p["role_assignments_detected"] = True
                    add_role_assignment(p, fixed_role, person, u, member_invested)
            for t in _split_tasks(week_task):
                if t not in p["tasks"]:
                    p["tasks"].append(t)

        out = []
        for p in projects.values():
            node_key = _match_node_key(p["node_text"])
            owners = [m for m in p["members"] if m["is_owner"]]
            unmatched = list(dict.fromkeys(m["name"] for m in p["members"] if not m["matched"]))
            warns = []
            if not node_key:
                warns.append(f"关键节点「{p['node_text']}」无法识别，需人工指定")
            if len(owners) > 1:
                warns.append("识别到多个负责人，请确认")
            if unmatched:
                warns.append(f"成员未匹配到用户：{','.join(unmatched)}")
            for member in p["members"]:
                member.pop("_key", None)
            for entries in p["role_assignments"].values():
                for entry in entries:
                    entry.pop("_key", None)
            out.append({**p, "current_node_key": node_key, "warnings": warns})
        return {"projects": out, "warnings": warnings, "errors": errors,
                "week_start": self.ps.week_start_of(date.today())}

    # ---------- confirm ----------
    def confirm(self, projects: list[dict], operator_id: int) -> dict:
        # 默认模板节点
        tpl = self.db.execute(select(TrTemplate).where(TrTemplate.status == "active").order_by(TrTemplate.id)).scalars().first()
        tpl_nodes = {n.node_key: n for n in self.db.execute(
            select(TrTemplateNode).where(TrTemplateNode.template_id == tpl.id).order_by(TrTemplateNode.sequence)).scalars().all()} if tpl else {}

        created, updated, failed = 0, 0, []
        for p in projects:
            try:
                self._import_one(p, tpl_nodes, operator_id)
                created += 1
            except Exception as e:  # noqa: BLE001
                failed.append({"project": p.get("name"), "error": str(e)[:200]})
        return {"created": created, "failed": failed}

    def _import_one(self, p: dict, tpl_nodes: dict, operator_id: int) -> None:
        # 找负责人
        owner = next((m for m in p["members"] if m.get("is_owner") and m.get("user_id")), None)
        if not owner:
            owner = next((m for m in p["members"] if m.get("user_id")), None)
        owner_id = owner["user_id"] if owner else operator_id

        # 项目（存在则更新）
        code = p.get("code") or f"P{date.today().year}{abs(hash(p['name'])) % 10000:04d}"
        proj = self.db.execute(select(Project).where(Project.code == code, Project.is_deleted.is_(False))).scalar_one_or_none()
        if not proj:
            proj = Project(name=p["name"], code=code, machine_model=p.get("machine_model"),
                           owner_id=owner_id, status="in_progress", created_by=operator_id)
            self.db.add(proj)
            self.db.flush()
        else:
            proj.name = p["name"]
            proj.machine_model = p.get("machine_model")

        # 成员（按 user_id 去重，避免 preview 重复项触发唯一约束；写入后 flush 保证后续 upsert 可见）
        seen_users = set()
        for m in p["members"]:
            m_uid = m.get("user_id")
            if not m_uid or m_uid in seen_users:
                continue
            seen_users.add(m_uid)
            role = "负责人" if (m.get("is_owner") or m_uid == owner_id) else (m.get("project_role") or "成员")
            member = self.db.execute(
                select(ProjectMember).where(
                    ProjectMember.project_id == proj.id, ProjectMember.user_id == m_uid)
            ).scalar_one_or_none()
            invested = m.get("is_invested", True)
            if member:
                member.is_deleted = False
                member.project_role = role
                member.is_invested = invested
            else:
                self.db.add(ProjectMember(project_id=proj.id, user_id=m_uid, project_role=role,
                                          is_invested=invested, joined_at=date.today()))
        self.db.flush()
        proj.owner_id = owner_id

        if p.get("role_assignments_detected"):
            role_ids = {}
            for raw_role, entries in (p.get("role_assignments") or {}).items():
                role = canonical_project_role(raw_role)
                if not role:
                    continue
                ids = []
                for entry in entries or []:
                    user_id = entry.get("user_id") if isinstance(entry, dict) else entry
                    if user_id and int(user_id) not in ids:
                        ids.append(int(user_id))
                role_ids[role] = ids
            ProjectService(self.db).replace_role_assignments(proj.id, role_ids)

        # 节点：按模板全量实例化，当前节点置为匹配到的
        current_key = p.get("current_node_key")
        if tpl_nodes and not self.db.execute(select(ProjectNode).where(ProjectNode.project_id == proj.id)).scalars().first():
            seqs = sorted(tpl_nodes.values(), key=lambda n: n.sequence)
            current_id = None
            reached = False
            for tn in seqs:
                is_current = (tn.node_key == current_key)
                status = "in_progress" if is_current else ("passed" if (current_key and not reached and tn.node_key != current_key and list(seqs).index(tn) < [x.node_key for x in seqs].index(current_key)) else "not_started")
                node = ProjectNode(project_id=proj.id, template_node_id=tn.id, node_key=tn.node_key,
                                   name=tn.name, sequence=tn.sequence, status=status)
                self.db.add(node)
                self.db.flush()
                if is_current:
                    current_id = node.id
                    reached = True
            proj.current_node_id = current_id or (seqs[0] and self.db.execute(
                select(ProjectNode).where(ProjectNode.project_id == proj.id).order_by(ProjectNode.sequence)).scalars().first().id)

        # 周目标（当周）
        if p.get("weekly_goal"):
            self.ps.set_weekly_goal(proj.id, date.today(), p["weekly_goal"], {"user_id": operator_id, "role": "admin"})

        # 任务（挂当前节点）
        if proj.current_node_id:
            for title in p.get("tasks", []):
                self.db.add(Task(project_node_id=proj.current_node_id, project_id=proj.id,
                                 title=title, status="in_progress", created_by=operator_id))
        self.db.commit()
