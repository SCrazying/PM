"""汇总引擎：项目周报 / 组内周报（按项目、按人）。"""
from datetime import date, timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.project_roles import PROJECT_ROLE_NAMES, canonical_project_role
from app.models.misc import Progress, ProjectWeeklyGoal, WeeklyGoalItem
from app.models.project import ACTIVE_PROJECT_STATUSES, Project, ProjectMember, ProjectNode, ProjectRoleAssignment, Task
from app.models.user import User
from app.services.progress_service import ProgressService


class ReportService:
    def __init__(self, db: Session):
        self.db = db
        self.ps = ProgressService(db)

    def _week_range(self, week_start: date) -> tuple[date, date]:
        ws = self.ps.week_start_of(week_start)
        return ws, ws + timedelta(days=6)

    def _project_role_summary(self, project: Project, members: list[tuple[ProjectMember | None, str]]) -> str:
        """生成可直接放入 Excel 单元格的项目角色摘要。"""
        role_names = {role: [] for role in PROJECT_ROLE_NAMES}
        assignments = self.db.execute(
            select(ProjectRoleAssignment, User.display_name)
            .join(User, User.id == ProjectRoleAssignment.user_id)
            .where(ProjectRoleAssignment.project_id == project.id)
        ).all()
        for assignment, display_name in assignments:
            if display_name and display_name not in role_names[assignment.role]:
                role_names[assignment.role].append(display_name)

        # 兼容尚未迁移角色分配记录的历史项目。
        for member, display_name in members:
            if not member:
                continue
            role = canonical_project_role(member.project_role)
            if role and display_name not in role_names[role]:
                role_names[role].append(display_name)

        lines = []
        owner = self.db.get(User, project.owner_id)
        if owner and owner.display_name:
            lines.append(f"负责人: {owner.display_name}")
        for role in PROJECT_ROLE_NAMES:
            if role_names[role]:
                lines.append(f"{role}: {'、'.join(role_names[role])}")
        return "\n".join(lines)

    # ---------- 项目周报 ----------
    def project_weekly(self, project_id: int, week_start: date) -> dict:
        ws, we = self._week_range(week_start)
        project = self.db.get(Project, project_id)
        project_members = self.db.execute(
            select(ProjectMember, User.display_name)
            .join(User, User.id == ProjectMember.user_id)
            .where(ProjectMember.project_id == project_id, ProjectMember.is_deleted.is_(False))
            .order_by(ProjectMember.id)
        ).all()
        role_text = self._project_role_summary(project, project_members)

        # 周目标
        goal_row = self.db.execute(
            select(ProjectWeeklyGoal).where(
                ProjectWeeklyGoal.project_id == project_id,
                ProjectWeeklyGoal.week_start == ws,
                ProjectWeeklyGoal.is_deleted.is_(False),
            )
        ).scalar_one_or_none()

        # 本周任务（计划区间与本周相交 OR 本周内完成）
        tasks = self.db.execute(
            select(Task).where(Task.project_id == project_id, Task.is_deleted.is_(False))
        ).scalars().all()
        week_tasks = []
        for t in tasks:
            in_plan = t.planned_start and t.planned_end and (t.planned_start <= we and t.planned_end >= ws)
            done_this_week = t.actual_end and (ws <= t.actual_end <= we)
            if in_plan or done_this_week or (t.status != "done" and t.planned_end is None):
                week_tasks.append({
                    "id": t.id, "title": t.title, "status": t.status,
                    "planned_start": t.planned_start, "planned_end": t.planned_end,
                    "actual_end": t.actual_end,
                    "overdue": bool(t.status != "done" and t.planned_end and t.planned_end < date.today()),
                    "assignee_id": t.assignee_id,
                })

        # 每日明细（按日期→人）
        progresses = self.db.execute(
            select(Progress, User.display_name).join(User, User.id == Progress.author_id).where(
                Progress.project_id == project_id,
                Progress.progress_date >= ws, Progress.progress_date <= we,
                Progress.is_deleted.is_(False),
            ).order_by(Progress.progress_date, Progress.id)
        ).all()
        daily = {}
        risks = []
        for p, uname in progresses:
            d = p.progress_date.isoformat()
            daily.setdefault(d, []).append({
                "author": uname, "today_work": p.today_work,
                "tomorrow_plan": p.tomorrow_plan, "risk": p.risk, "risk_resolved": p.risk_resolved,
            })
            if p.risk:
                risks.append({"progress_id": p.id, "date": d, "author": uname, "risk": p.risk,
                              "resolved": p.risk_resolved})

        # 全部一级节点（周会"当前节点"列显示所有节点，每行一个）
        top_nodes = self.db.execute(
            select(ProjectNode).where(
                ProjectNode.project_id == project_id, ProjectNode.parent_id.is_(None),
                ProjectNode.is_deleted.is_(False),
            ).order_by(ProjectNode.sequence)
        ).scalars().all()
        nodes = [{
            "id": n.id, "node_key": n.node_key, "name": n.name, "status": n.status,
            "planned_end": n.planned_end,
            "overdue": bool(n.status != "passed" and n.planned_end and n.planned_end < date.today()),
            "is_current": n.id == project.current_node_id,
        } for n in top_nodes]

        # 当前节点 + 全部节点子节点（周会视图按节点分组展示全部子节点）
        current_node = None
        subnodes = []
        if project.current_node_id:
            node = self.db.get(ProjectNode, project.current_node_id)
            if node:
                current_node = {
                    "id": node.id, "node_key": node.node_key, "name": node.name,
                    "planned_start": node.planned_start, "planned_end": node.planned_end,
                    "overdue": bool(node.status != "passed" and node.planned_end and node.planned_end < date.today()),
                }
        # 全部顶层节点及其子节点（仅含有子节点的节点）
        node_subnodes = []
        top_ids = [n.id for n in top_nodes]
        if top_ids:
            sub_rows = self.db.execute(
                select(ProjectNode).where(
                    ProjectNode.parent_id.in_(top_ids), ProjectNode.is_deleted.is_(False),
                ).order_by(ProjectNode.sequence)
            ).scalars().all()
            subs_by_parent: dict[int, list] = {}
            for s in sub_rows:
                subs_by_parent.setdefault(s.parent_id, []).append(s)
            for n in top_nodes:
                subs = subs_by_parent.get(n.id)
                if not subs:
                    continue
                node_subnodes.append({
                    "node_id": n.id, "node_key": n.node_key, "name": n.name,
                    "subnodes": [
                        {"id": s.id, "name": s.name, "status": s.status, "planned_end": s.planned_end,
                         "actual_end": s.actual_end,
                         "overdue": bool(s.status != "done" and s.planned_end and s.planned_end < date.today())}
                        for s in subs
                    ],
                })
                if n.id == project.current_node_id:
                    subnodes = node_subnodes[-1]["subnodes"]

        # 周目标条目（本周）
        goal_rows = self.db.execute(
            select(WeeklyGoalItem).where(
                WeeklyGoalItem.project_id == project_id, WeeklyGoalItem.week_start == ws,
            ).order_by(WeeklyGoalItem.sequence, WeeklyGoalItem.id)
        ).scalars().all()
        goal_user_ids = {g.user_id for g in goal_rows if g.user_id}
        goal_names = {}
        if goal_user_ids:
            goal_names = {u.id: u.display_name for u in self.db.execute(
                select(User).where(User.id.in_(goal_user_ids))).scalars().all()}
        goal_items = [{
            "id": g.id, "goal": g.goal, "deadline": g.deadline,
            "done": g.done, "done_at": g.done_at,
            "user_id": g.user_id, "user_name": goal_names.get(g.user_id),
            "overdue": bool(not g.done and g.deadline and g.deadline < date.today()),
        } for g in goal_rows]

        return {
            "project": {"id": project.id, "name": project.name, "code": project.code,
                        "machine_model": project.machine_model, "health": project.health,
                        "status": project.status, "description": project.description,
                        "current_node": current_node,
                        "nodes": nodes, "project_roles": role_text},
            "week_start": ws, "week_end": we,
            "weekly_goal": goal_row.goal if goal_row else None,
            "weekly_goal_items": goal_items,
            "tasks": week_tasks,
            "daily": daily,
            "risks": risks,
            "subnodes": subnodes,
            "node_subnodes": node_subnodes,
        }

    # ---------- 组内周报：按项目 ----------
    def group_weekly_by_project(self, week_start: date) -> list[dict]:
        projects = self.db.execute(
            select(Project).where(Project.is_deleted.is_(False), Project.status.in_(ACTIVE_PROJECT_STATUSES))
            .order_by(Project.id)
        ).scalars().all()
        return [self.project_weekly(p.id, week_start)["project"] | {"week_start": self._week_range(week_start)[0],
                "weekly_goal": self.project_weekly(p.id, week_start)["weekly_goal"]} for p in projects]

    # ---------- 项目台账 Excel 导出（M5） ----------
    def export_ledger_xlsx(self, week_start: date, scope: str = "weekly") -> BytesIO:
        """导出项目台账 Excel。
        scope=weekly：本周台账，与周会视图「按项目」视图列结构一致（每项目一行）；
        scope=all：项目台账（每周任务合集），保持按成员一行的历史格式。"""
        ws_start, ws_end = self._week_range(week_start)
        wb = Workbook()
        ws = wb.active
        header_fill = PatternFill("solid", fgColor="4F6EF7")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin = Side(style="thin", color="D9E0EA")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def _header(headers, widths, title):
            ws.title = title
            for col, (header, width) in enumerate(zip(headers, widths), start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
                ws.column_dimensions[get_column_letter(col)].width = width
            ws.row_dimensions[1].height = 28
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        projects = self.db.execute(
            select(Project).where(Project.is_deleted.is_(False)).order_by(Project.name)
        ).scalars().all()

        # ---------- 本周台账：与周会视图「按项目」视图一致，每项目一行 ----------
        if scope == "weekly":
            status_labels = {"not_started": "未开始", "in_progress": "进行中", "delayed": "延期",
                             "completed": "已完成", "suspended": "暂停"}
            headers = ["机型", "项目名称", "项目描述", "状态", "项目角色", "当前节点", "子节点", "周目标", "每日进展"]
            widths = [14, 22, 26, 10, 24, 26, 30, 32, 50]
            _header(headers, widths, "本周台账")

            def short(d):
                return d.isoformat()[5:] if d else ""

            def fmt_nodes(nodes):
                """当前节点列：未通过的一级节点，TR3 08-06 超期"""
                lines = []
                for n in nodes:
                    if n["status"] == "passed":
                        continue
                    end = short(n["planned_end"]) or "—"
                    lines.append(f"{n['node_key']} {end}{' 超期' if n['overdue'] else ''}")
                return "\n".join(lines) or "未设置"

            def fmt_subnodes(subnodes):
                """子节点列：✓ 已完成 / ○ 待办，含日期与延期标记"""
                lines = []
                for s in subnodes:
                    if s["status"] == "done":
                        lines.append(f"✓ {short(s['actual_end'])} {s['name']}".strip())
                    elif s["overdue"]:
                        lines.append(f"○ {short(s['planned_end'])} 延期 {s['name']}")
                    elif s["planned_end"]:
                        lines.append(f"○ {short(s['planned_end'])} {s['name']}")
                    else:
                        lines.append(f"○ {s['name']}")
                return "\n".join(lines) or "无"

            def fmt_goals(items, legacy):
                """周目标列：✓ 已完成条目（含负责人）；无条目时回落历史自由文本"""
                if items:
                    lines = []
                    for g in items:
                        name = f"{g['user_name']}：" if g.get("user_name") else ""
                        if g["done"]:
                            lines.append(f"✓ {g['done_at'] or ''} {name}{g['goal']}".strip())
                        elif g["overdue"]:
                            lines.append(f"○ {name}{g['goal']} 超期 {g['deadline']}")
                        elif g["deadline"]:
                            lines.append(f"○ {name}{g['goal']} {g['deadline']}")
                        else:
                            lines.append(f"○ {name}{g['goal']}")
                    return "\n".join(lines)
                return legacy or "（未设周目标）"

            def fmt_daily(daily):
                """每日进展列：日期 姓名：进展（风险），与周会视图一致"""
                lines = []
                for d in sorted(daily.keys(), reverse=True):
                    for it in daily[d]:
                        risk = "（风险）" if it["risk"] and not it["risk_resolved"] else ""
                        lines.append(f"{d[5:]} {it['author']}：{it['today_work']}{risk}")
                return "\n".join(lines) or "无"

            row_no = 2
            for project in projects:
                data = self.project_weekly(project.id, ws_start)
                prj = data["project"]
                values = [
                    prj["machine_model"] or "",
                    prj["name"],
                    prj["description"] or "",
                    status_labels.get(prj["status"], prj["status"]),
                    prj["project_roles"] or "",
                    fmt_nodes(prj["nodes"]),
                    fmt_subnodes(data["subnodes"]),
                    fmt_goals(data["weekly_goal_items"], data["weekly_goal"]),
                    fmt_daily(data["daily"]),
                ]
                for col, value in enumerate(values, start=1):
                    cell = ws.cell(row=row_no, column=col, value=value)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    cell.border = border
                line_count = max((str(v).count("\n") + 1 for v in values), default=1)
                ws.row_dimensions[row_no].height = min(180, max(42, 16 * line_count))
                row_no += 1
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output

        # ---------- 项目台账（每周任务合集）：按成员一行 ----------
        ws.title = "项目台账"
        task_header = "项目任务"
        headers = ["机型", "项目", "是否投入", "项目角色", "关键节点", "周目标", task_header]
        widths = [14, 24, 12, 22, 18, 30, 52]
        _header(headers, widths, "项目台账")
        row_no = 2
        for project in projects:
            current = self.db.get(ProjectNode, project.current_node_id) if project.current_node_id else None
            goal = self.db.execute(
                select(ProjectWeeklyGoal).where(
                    ProjectWeeklyGoal.project_id == project.id,
                    ProjectWeeklyGoal.is_deleted.is_(False),
                ).order_by(ProjectWeeklyGoal.week_start.desc())
            ).scalars().first()
            members = self.db.execute(
                select(ProjectMember, User.display_name).join(User, User.id == ProjectMember.user_id).where(
                    ProjectMember.project_id == project.id,
                    ProjectMember.is_deleted.is_(False),
                ).order_by(ProjectMember.id)
            ).all()
            # 没有成员时仍保留项目一行，方便发现数据缺口
            if not members:
                members = [(None, "未分配")]
            role_text = self._project_role_summary(project, members)
            for member, display_name in members:
                member_tasks = []
                member_progress = []
                if member:
                    tasks = self.db.execute(select(Task).where(
                        Task.project_id == project.id, Task.assignee_id == member.user_id,
                        Task.is_deleted.is_(False),
                    )).scalars().all()
                    member_tasks = [f"{'✓' if t.status == 'done' else '○'} {t.title}" for t in tasks]
                    progresses = self.db.execute(
                        select(Progress).where(
                            Progress.project_id == project.id, Progress.author_id == member.user_id,
                            Progress.is_deleted.is_(False),
                        ).order_by(Progress.progress_date)
                    ).scalars().all()
                    member_progress = [f"[{p.progress_date}] {p.today_work}" for p in progresses]
                task_text = "\n".join(member_tasks + member_progress)
                values = [
                    project.machine_model or "",
                    project.name,
                    "是" if member and member.is_invested else "否",
                    role_text,
                    f"{current.node_key} {current.name}" if current else "",
                    goal.goal if goal else "",
                    task_text,
                ]
                for col, value in enumerate(values, start=1):
                    cell = ws.cell(row=row_no, column=col, value=value)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    cell.border = border
                role_lines = role_text.count("\n") + 1 if role_text else 1
                task_lines = task_text.count("\n") + 1 if task_text else 1
                ws.row_dimensions[row_no].height = min(180, max(42, 18 * max(role_lines, task_lines)))
                row_no += 1
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def export_completion_xlsx(self) -> BytesIO:
        """项目完成台账：每项目一行，含各 TR 节点完成状态与项目完成度。"""
        wb = Workbook()
        ws = wb.active
        ws.title = "项目完成台账"
        headers = ["机型", "项目", "负责人", "当前节点", "节点完成", "TR节点状态", "项目完成度"]
        widths = [14, 26, 12, 18, 12, 46, 12]
        header_fill = PatternFill("solid", fgColor="4F6EF7")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin = Side(style="thin", color="D9E0EA")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col, (header, width) in enumerate(zip(headers, widths), start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill; cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        today = date.today()
        projects = self.db.execute(
            select(Project).where(Project.is_deleted.is_(False)).order_by(Project.name)
        ).scalars().all()
        row_no = 2
        for p in projects:
            owner = self.db.get(User, p.owner_id)
            current = self.db.get(ProjectNode, p.current_node_id) if p.current_node_id else None
            nodes = list(self.db.execute(
                select(ProjectNode).where(
                    ProjectNode.project_id == p.id, ProjectNode.parent_id.is_(None),
                    ProjectNode.is_deleted.is_(False),
                ).order_by(ProjectNode.sequence)
            ).scalars().all())
            passed = sum(1 for n in nodes if n.status == "passed")
            total = len(nodes)
            status_parts = []
            for n in nodes:
                mark = "✓" if n.status == "passed" else "→" if n.status == "in_progress" else "○"
                if n.status != "passed" and n.planned_end and n.planned_end < today:
                    mark = "⚠"
                status_parts.append(f"{n.node_key}{mark}")
            values = [
                p.machine_model or "",
                p.name,
                owner.display_name if owner else "",
                f"{current.node_key} {current.name}" if current else "",
                f"{passed}/{total}",
                " ".join(status_parts),
                f"{round(passed / total * 100)}%" if total else "0%",
            ]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row_no, column=col, value=value)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = border
            ws.row_dimensions[row_no].height = 24
            row_no += 1
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # ---------- 组内周报：按人 ----------
    def group_weekly_by_person(self, week_start: date) -> list[dict]:
        ws, we = self._week_range(week_start)
        users = self.db.execute(select(User).where(User.status == "active").order_by(User.id)).scalars().all()
        out = []
        for u in users:
            # 参与的项目（含角色/投入）
            mems = self.db.execute(
                select(ProjectMember, Project).join(Project, Project.id == ProjectMember.project_id).where(
                    ProjectMember.user_id == u.id, ProjectMember.is_deleted.is_(False),
                    Project.is_deleted.is_(False), Project.status.in_(ACTIVE_PROJECT_STATUSES),
                )
            ).all()
            if not mems:
                continue
            projects = []
            for mem, proj in mems:
                # 本周此人在此项目的进展条数
                cnt = self.db.execute(
                    select(Progress).where(
                        Progress.project_id == proj.id, Progress.author_id == u.id,
                        Progress.progress_date >= ws, Progress.progress_date <= we,
                        Progress.is_deleted.is_(False),
                    )
                ).scalars().all()
                projects.append({
                    "project_id": proj.id, "name": proj.name, "code": proj.code,
                    "project_role": mem.project_role, "is_invested": mem.is_invested,
                    "progress_count": len(cnt),
                })
            # 本周进展明细
            progresses = self.db.execute(
                select(Progress, Project.name).join(Project, Project.id == Progress.project_id).where(
                    Progress.author_id == u.id, Progress.progress_date >= ws, Progress.progress_date <= we,
                    Progress.is_deleted.is_(False),
                ).order_by(Progress.progress_date)
            ).all()
            daily = {}
            for p, pname in progresses:
                daily.setdefault(p.progress_date.isoformat(), []).append(
                    {"project": pname, "today_work": p.today_work, "risk": p.risk})
            out.append({"user_id": u.id, "display_name": u.display_name,
                        "projects": projects, "daily": daily})
        return out
